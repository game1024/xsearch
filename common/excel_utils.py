from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


def style_heading(excel_file, font_name='黑体', font_size=14, horizontal='center', vertical='center', sheet_name=None):
    wb = load_workbook(excel_file)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    font = Font(name=font_name, size=font_size, bold=True)
    alignment = Alignment(horizontal=horizontal, vertical=vertical)
    for cell in ws[1]:
        cell.font = font
        cell.alignment = alignment